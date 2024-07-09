from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
import sqlite3
import os
import random
from pydub import AudioSegment
from datetime import datetime
import time
from pydantic import BaseModel
from openpyxl import load_workbook
import sys
import yt_dlp

class Dancer(BaseModel):
    name: str

class ExcelPlaylist(BaseModel):
    countdown: bool
    countdownCrossfade: bool
    intro: bool
    outro: bool
    preTime: int
    postTime: int
    fadeInTime: int
    fadeOutTime: int
    countdownVoice: str
    countdownLength: int
    coverImage: str
    playlistAmount: int
    randomizePlaylist: bool
    backendConformation: bool
    tenSecondSilenceAtEnd: bool
    fileName: str
    checkYTURL: bool
    removeDancer: bool
    checkYTURLPosition: int


app = FastAPI()

origins = [
    "http://localhost:3000",
    "http://localhost:8080",
    "http://localhost:8081",
    "http://127.0.0.1:8080",
    "http://192.168.0.165:8080/",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

monsta_x_love = {
        "URL": "https://youtu.be/wssbMRrXRD8",
        "Title": "love",
        "Artist": "monstax",
        "Start": 84,
        "End": 91,
        "Description": "Custom: Intro",
        "Dancer": "None"
    }

bts_run = {
        "URL": "https://youtu.be/2WBwJD6hldA",
        "Title": "run",
        "Artist": "bts",
        "Start": 228,
        "End": 233,
        "Description": "Custom: Outro",
        "Dancer": "None"
    }

def match_target_amplitude(sound, target_dBFS):
    change_in_dBFS = target_dBFS - sound.dBFS
    return sound.apply_gain(change_in_dBFS)

def remove_special_char_and_lower(string):
    return ''.join(e for e in str(string) if e.isalnum()).lower()

def shuffle_playlist(playlist):
    playlist_temp = playlist
    no_same_artist_next_to_each_other = False
    random.shuffle(playlist_temp)
    while no_same_artist_next_to_each_other == False:
        count = 0
        for x in range(len(playlist)):
            if count > 0:
                if playlist[x]["Artist"] == playlist[x-1]["Artist"]:
                    print("2 Songs with the same Artist Back-to-Back! Reroll...")
                    random.shuffle(playlist_temp)
                    break
            else:
                count += 1
        else:
            no_same_artist_next_to_each_other = True
    return playlist_temp

def create_playlist(songlist, intro, outro, countdownLength, countdownVoice, countdownCrossfade, tenSecondSilenceAtEnd, coverImage, preTime, postTime, fadeInTime, fadeOutTime, countdown, fileName, removeDancer):
    if(intro):
        songlist.insert(0, monsta_x_love)

    if(outro):
        songlist.append(bts_run)

    # Download mp4 files from YouTube
    raw_folder_content = os.listdir("raw")
    for song in songlist:
        file_name = song["ArtistFile"] + " - " + song["TitleFile"] + ".m4a"

        if file_name not in raw_folder_content:
            print(f"{file_name} is missing! Downloading... ")

            # Old: PyTube  
            """ yt = YouTube(song["URL"])
            audio_stream = yt.streams.get_audio_only()
            audio_stream.download("raw/", file_name) """

            URLS = [song["URL"]]

            ydl_opts = {
                'format': 'm4a/bestaudio/best',
                # ℹ️ See help(yt_dlp.postprocessor) for a list of available Postprocessors and their arguments
                'postprocessors': [{  # Extract audio using ffmpeg
                    'key': 'FFmpegExtractAudio',
                    'preferredcodec': 'm4a',
                }],
                'outtmpl': f'raw/{file_name}'
            }
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                error_code = ydl.download(URLS)
            
        
    print("Download missing songs complete!")

    if (countdown and not intro):
        chapters = [["00:00:00", 
            0, 
            "RDG Start",
            "",
            ""]]
    else:
        chapters = []

    song_counter = 0

    if countdownLength == 3:
        countdown = AudioSegment.from_file(f"templates/countdown/{countdownVoice}/three.mp3")
    elif countdownLength == 5:
        countdown = AudioSegment.from_file(f"templates/countdown/{countdownVoice}/five.mp3")
    else:
        countdown = AudioSegment.from_file(f"templates/countdown/{countdownVoice}/three.mp3")

    countdown_dancebreak = AudioSegment.from_file(f"templates/countdown/{countdownVoice}/dancebreak.mp3")
    countdown_blue = AudioSegment.from_file(f"templates/countdown/{countdownVoice}/blue.mp3")
    countdown_red = AudioSegment.from_file(f"templates/countdown/{countdownVoice}/red.mp3")
    countdown_cool = AudioSegment.from_file(f"templates/countdown/{countdownVoice}/cool.mp3")
    countdown_warm = AudioSegment.from_file(f"templates/countdown/{countdownVoice}/warm.mp3")
    countdown_summer = AudioSegment.from_file(f"templates/countdown/{countdownVoice}/summer.mp3")
    export = AudioSegment.empty()

    print("Combining the songs...")
    for song in songlist:

        if (countdown and not (song["Description"] == "Custom: Intro" or song["Description"] == "Custom: Outro")):
            if (song["Dancer"] != None):
                if (song["Dancer"].lower() == "blue"):
                    export += countdown_blue
                if (song["Dancer"].lower() == "red"):
                    export += countdown_red
                if (song["Dancer"].lower() == "cool"):
                    export += countdown_cool
                if (song["Dancer"].lower() == "warm"):
                    export += countdown_warm
                if (song["Dancer"].lower() == "summer"):
                    export += countdown_summer
            
            if (song["Description"] != None):
                if (''.join(e for e in song["Description"] if e.isalnum()).lower() == "dancebreak"):
                    export += countdown_dancebreak

            export += countdown

        # Chapters
        song_counter += 1
        timestamp = time.strftime('%H:%M:%S', time.gmtime(int(export.duration_seconds)))
        chapters.append([
            timestamp, 
            song_counter, 
            " - ".join([song["Artist"], song["Title"]]),
            f"({song['Description']})",
            f"({song['Dancer']})"
            ])

        # Audio
        file_name = song["ArtistFile"] + " - " + song["TitleFile"] + ".m4a"
        song_snippet = AudioSegment.from_file("raw/" + file_name)[(song["Start"] - preTime) * 1000:(song["End"] + postTime) * 1000].fade_in(1000 * fadeInTime).fade_out(1000 * fadeOutTime)
        song_snippet_normalized = match_target_amplitude(song_snippet, -10.0)
        if countdown and countdownCrossfade and not (song["Description"] == "Custom: Intro" or song["Description"] == "Custom: Outro"):
            export = export.append(song_snippet_normalized, crossfade=1000)
        else:
            export += song_snippet_normalized

    #Add 10 Seconds Silence at the end
    if tenSecondSilenceAtEnd == True:
        song_counter += 1
        timestamp = time.strftime('%H:%M:%S', time.gmtime(int(export.duration_seconds)))
        chapters.append([timestamp, song_counter, "10 Second Cooldown", "", ""])
        
        export += AudioSegment.silent(duration=10000)
    
    if "NoFileName" in fileName:
        export_file_name = f"{''.join([str(datetime.now().strftime('%Y-%m-%d_%H_%M_%S'))])}{fileName}.mp3"
    else:
        export_file_name = fileName + ".mp3"

    #Entfernt Leerzeichen im String, ansonsten funktioniert der MMPEG Export nicht
    export_file_name = export_file_name.replace(" ", "")

    print("Exporting the MP3...")
    export.export("export/" + export_file_name , format="mp3")

    chapter_summary_comment = ""
    chapter_summary_YouTube = ""

    if removeDancer:
        for chapter in chapters:
            chapter_summary_comment += f"{chapter[0]} {chapter[1]} {chapter[2]} {chapter[3]}newLine"
            chapter_summary_YouTube += f"{chapter[0]} {chapter[1]} {chapter[2]} {chapter[3]}\n"
    else:
        for chapter in chapters:
            chapter_summary_comment += f"{chapter[0]} {chapter[1]} {chapter[2]} {chapter[3]} {chapter[4]}newLine"
            chapter_summary_YouTube += f"{chapter[0]} {chapter[1]} {chapter[2]} {chapter[3]} {chapter[4]}\n"

    print("Converting the MP3 to MP4...")

    os.system(f"""ffmpeg.exe -loop 1 -framerate 1 -i templates/image/{coverImage}.jpg -i export/{export_file_name} -map 0:v -map 1:a -r 10 -vf \"scale='iw-mod(iw,2)\':\'ih-mod(ih,2)\',format=yuv420p\" -movflags +faststart -shortest -fflags +shortest -max_interleave_delta 100M -metadata comment=\"{chapter_summary_comment}\" export/{export_file_name}.mp4""")

    print("Playlist Creation completed!")
    return chapter_summary_YouTube

@app.get("/")
def read_root():
    return {"Data": "Hello World"}

@app.post("/createExcelPlaylist")
def create_item(playlist: ExcelPlaylist):

    wb = load_workbook(filename="songlist.xlsx", data_only=True)
    sheet = wb["Songlist"]

    song_list = []

    for x in range(2, sheet.max_row + 1):
        startTime = sheet.cell(x, 10).value
        if(sheet.cell(x, 10).value < playlist.preTime):
            startTime = playlist.preTime
        
        # Kontrolliert, ob die Startzeit vor der Endzeit ist
        if sheet.cell(x, 11).value <= startTime:
            print(f"Fehler bei {sheet.cell(x, 3).value} (Position {x}): Startzeit ist größer als die Endzeit")
            return f"Fehler bei '{sheet.cell(x, 3).value}': Startzeit ({startTime}) ist größer als die Endzeit ({sheet.cell(x, 11).value})"

        # Wandelt Titel von int zu String um (Bsp: 458 von CIX)
        title = str(sheet.cell(x, 3).value)

        song_list.append({
            "URL": sheet.cell(x, 1).value,
            "Artist": sheet.cell(x, 2).value,
            "ArtistFile": remove_special_char_and_lower(sheet.cell(x, 2).value),
            "Title": title,
            "TitleFile": remove_special_char_and_lower(sheet.cell(x, 3).value),
            "Description": sheet.cell(x, 4).value,
            "Dancer": sheet.cell(x, 5).value,
            "Start": startTime,
            "End": sheet.cell(x, 11).value
        })

    wb.close()

    # Kontrolliert, ob alle YT URLs gültig sind
    if playlist.checkYTURL:
        count = 0
        for yt in song_list:
            count += 1
            print(f"Testing URL of -{yt['Title']}-")
            if count < playlist.checkYTURLPosition:
                print("URL is already tested and will be skipped...")
            else: 
                search_title = ''.join(e for e in str(yt["Title"]) if e.isalnum()).lower()
                
                # Old
                """ r = requests.get(yt["URL"])
                if "unavailable/unavailable_video.png" in r.text:
                    return f"-{yt['Title']}- deren URL funktioniert nicht\n{count}" """
                
                # ℹ️ See help(yt_dlp.YoutubeDL) for a list of available options and public functions
                ydl_opts = {}
                with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                    info = ydl.extract_info(yt["URL"], download=False)

                    # ℹ️ ydl.sanitize_info makes the info json-serializable
                    search_string = ''.join(e for e in str(info['title']) if e.isalnum()).lower()

                if search_title not in search_string:
                    return f"-{yt['Title']}- hat die falsche URL (falscher Song?)\n{''.join(e for e in yt['Title'] if e.isalnum()).lower()}\n{count}\n{info['title']}n{''.join(e for e in info['title'] if e.isalnum()).lower()}"

    playlist_list = [[] for i in range(playlist.playlistAmount)]

    yt_chapter_summary = ""

    song_list_2 = []

    # Putting specific Songs into specific Playlist
    for song in song_list:
        if song["Dancer"] is not None:
            dancer_name = ''.join(e for e in song["Dancer"] if e.isalnum()).lower()
            if "playlist" in dancer_name:
                print(f"Special instruction: \"{song['Title']}\" will be in Playlist {dancer_name[-1]}")
                playlist_list[int(dancer_name[-1]) - 1].append(song)
            else:
                song_list_2.append(song)
        else:
            song_list_2.append(song)

    # Distribute Songs into the amount of requested Playlist
    for song_2 in song_list_2:
        min(playlist_list, key=len).append(song_2)

    # Shuffle Playlist
    # Disabled, or "# Putting specific Songs into specific Playlist" won't work!
    # random.shuffle(playlist_list)

    for playlist_x in playlist_list:
        # Playlist Randomizer
        if playlist.randomizePlaylist:
            print("Randomizing the Playlist...")
            playlist_x = shuffle_playlist(playlist_x)

        # Randomizer Reroll Backend Conformation
        if playlist.backendConformation:
            answer = "r"
            while answer == "r":
                print("Proposal:")
                for x in playlist_x:
                    print(f"{x['Artist']} - {x['Title']} ({x['Description']}) ({x['Dancer']})")
                print("Any button: OK | r: Randomize again | n: Abort")
                answer = input()
                if answer == "n":
                    sys.exit("Operation aborted!")
                if answer == "r":
                    playlist_x = shuffle_playlist(playlist_x)

    

    if playlist.fileName == "":
        file_name = "NoFileName"
    else:
        file_name = playlist.fileName

    # Extra für Stuttgart: Letzte Playlist soll KEINEN 10-Sekunden Cooldown haben
    if playlist.coverImage == "RDGStuttgart":
        for n in range(len(playlist_list) - 1):
            yt_chapter_summary += f"Playlist {n+1}\n"
            yt_chapter_summary += create_playlist(playlist_list[n], playlist.intro, playlist.outro, playlist.countdownLength, playlist.countdownVoice, playlist.countdownCrossfade, playlist.tenSecondSilenceAtEnd, playlist.coverImage, playlist.preTime, playlist.postTime, playlist.fadeInTime, playlist.fadeOutTime, playlist.countdown, f"{file_name}_Playlist_{n + 1}", playlist.removeDancer)
            yt_chapter_summary += "\n"
        yt_chapter_summary += f"Playlist {len(playlist_list)}\n"
        yt_chapter_summary += create_playlist(playlist_list[len(playlist_list) - 1], playlist.intro, playlist.outro, playlist.countdownLength, playlist.countdownVoice, playlist.countdownCrossfade, False, playlist.coverImage, playlist.preTime, playlist.postTime, playlist.fadeInTime, playlist.fadeOutTime, playlist.countdown, f"{file_name}_Playlist_{len(playlist_list)}", playlist.removeDancer)
    else: # Für alle anderen RDGs
        for n in range(len(playlist_list)):
            yt_chapter_summary += f"Playlist {n+1}\n"
            yt_chapter_summary += create_playlist(playlist_list[n], playlist.intro, playlist.outro, playlist.countdownLength, playlist.countdownVoice, playlist.countdownCrossfade, playlist.tenSecondSilenceAtEnd, playlist.coverImage, playlist.preTime, playlist.postTime, playlist.fadeInTime, playlist.fadeOutTime, playlist.countdown, f"{file_name}_Playlist_{n + 1}", playlist.removeDancer)
            yt_chapter_summary += "\n"

    return yt_chapter_summary

@app.get("/showExcel")
def read_root():
    wb = load_workbook(filename="songlist.xlsx", data_only=True)
    sheet = wb["Songlist"]

    answer = []

    for x in range(2, sheet.max_row + 1):
        answer.append({
            "URL": sheet.cell(x, 1).value,
            "Artist": sheet.cell(x, 2).value,
            "Title": sheet.cell(x, 3).value,
            "Description": sheet.cell(x, 4).value,
            "Dancer": sheet.cell(x, 5).value,
            "Start": sheet.cell(x, 10).value,
            "End": sheet.cell(x, 11).value
        })
    
    wb.close()

    return answer

@app.post("/dancer/")
def create_item(dancer: Dancer):
    con = sqlite3.connect("songlist.db")
    cur = con.cursor()

    # res = cur.execute(f"""
    # INSERT INTO
    # DancerNew (Name)
    # VALUES ('{dancer.name}')
    # """)

    # res = cur.execute(f"""
    # CREATE TABLE DancerNew (
    #     DancerID INTEGER PRIMARY KEY,
    #     Name TEXT NOT NULL
    # )
    # """)

    # res = cur.execute(f"""
    # DROP TABLE DancerNew
    # """)

    # data = res.fetchall()
    con.close()
    return dancer

@app.get("/createVAOPlaylist/{dancer_ids}")
def read_item(dancer_ids):
    dancer_array = dancer_ids.split(",")
    condition_array = []
    for x in dancer_array:
        condition_array.append(f"dancer.DancerID={x}")
    condition_string =  " OR ".join(condition_array)

    con = sqlite3.connect("songlist.db")
    cur = con.cursor()
    res = cur.execute(f"""
    SELECT 
    song.URL, song.Title, song.Artist,
    sequence.Start, sequence.End, sequence.Description
    FROM sequence
    INNER JOIN song ON sequence.SongID=song.SongID
    INNER JOIN mastery ON sequence.SequenceID=mastery.SequenceID
    INNER JOIN dancer ON mastery.DancerID=dancer.DancerID
    WHERE {condition_string}
    """)
    data = res.fetchall()
    con.close()

    monsta_x_love = {
        "URL": "https://youtu.be/wssbMRrXRD8",
        "Title": "LOVE",
        "Artist": "MONSTA X",
        "Start": 84,
        "End": 91,
        "Description": "Custom: Intro"
    }

    bts_run = {
        "URL": "https://youtu.be/2WBwJD6hldA",
        "Title": "RUN",
        "Artist": "BTS",
        "Start": 228,
        "End": 233,
        "Description": "Custom: Outro" 
    }

    song_list = []

    for entry in data:
        song_list.append({
            "URL": entry[0],
            "Title": entry[1],
            "Artist": entry[2],
            "Start": entry[3],
            "End": entry[4],
            "Description": entry[5]
        })

    # Shuffle Playlist and reduce the length to 97 songs
    random.shuffle(song_list)
    song_list = song_list[:97]
    song_list.append(bts_run)
    song_list.insert(0, monsta_x_love)

    # Download mp4 files from YouTube
    raw_folder_content = os.listdir("raw")
    for song in song_list:
        file_name = song["Artist"] + " - " + song["Title"] + ".mp4"
        file_name = file_name.casefold()

        if file_name not in raw_folder_content:
            print(f"{file_name} is missing! Downloading... ")

            # Won't work because pytube is no more
            # yt = YouTube(song["URL"]) 
            # audio_stream = yt.streams.get_audio_only()
            # audio_stream.download("raw/", file_name)
        else:
            print(f"{file_name} is already downloaded.")

    chapters = []
    song_counter = 0
    export = AudioSegment.empty()

    print("Creating Export...")
    for song in song_list:

        # Chapters
        song_counter += 1
        timestamp = time.strftime('%H:%M:%S', time.gmtime(int(export.duration_seconds)))
        chapters.append([
            timestamp, 
            song_counter, 
            " - ".join([song["Artist"], song["Title"]]),
            f"({song['Description']})"
            ])

        # Audio
        file_name = song["Artist"] + " - " + song["Title"] + ".mp4"
        file_name = file_name.casefold()
        song_snippet = AudioSegment.from_file("raw/" + file_name)[(song["Start"] - 10) * 1000:(song["End"] + 2) * 1000].fade_in(2000).fade_out(2000)
        export += song_snippet

    export_file_name = "".join([str(datetime.now().strftime("%Y-%m-%d_%H_%M_%S"))]) + ".mp3"

    print("Exporting...")
    export.export("export/" + export_file_name , format="mp3")

    chapter_summary_comment = ""
    chapter_summary_YouTube = ""

    for chapter in chapters:
        chapter_summary_comment += f"{chapter[0]} {chapter[1]} {chapter[2]} {chapter[3]}newLine"
        chapter_summary_YouTube += f"{chapter[0]} {chapter[1]} {chapter[2]} {chapter[3]}\n"

    os.system(f"""ffmpeg.exe -loop 1 -framerate 1 -i image.jpg -i export/{export_file_name} -map 0:v -map 1:a -r 10 -vf \"scale='iw-mod(iw,2)\':\'ih-mod(ih,2)\',format=yuv420p\" -movflags +faststart -shortest -fflags +shortest -max_interleave_delta 100M -metadata comment=\"{chapter_summary_comment}\" export/{export_file_name}.mp4""")

    return {"Data": chapter_summary_YouTube}


@app.get("/showPreview/{dancer_ids}")
def read_item(dancer_ids):
    dancer_array = dancer_ids.split(",")
    condition_array = []
    for x in dancer_array:
        condition_array.append(f"dancer.DancerID={x}")
    condition_string =  " OR ".join(condition_array)

    con = sqlite3.connect("songlist.db")
    cur = con.cursor()
    res = cur.execute(f"""
    SELECT 
    song.URL, song.Title, song.Artist, 
    sequence.Description, sequence.Start, sequence.End, sequence.LastTimeInRDG, sequence.comment, 
    dancer.Name,
    mastery.MasteryID
    FROM sequence
    INNER JOIN song ON sequence.SongID=song.SongID
    INNER JOIN mastery ON sequence.SequenceID=mastery.SequenceID
    INNER JOIN dancer ON mastery.DancerID=dancer.DancerID
    WHERE {condition_string}
    """)
    data = res.fetchall()
    con.close()

    answer = []
    for entry in data:
        answer.append(
            {
                "Title": entry[1],
                "Artist": entry[2],
                "Description": entry[3],
                "Dancer": entry[8],
                "MasteryID": entry[9],
            }
        )
    return answer

@app.get("/mastery")
def read_root():
    con = sqlite3.connect("songlist.db")
    cur = con.cursor()
    res = cur.execute("""
    SELECT 
    mastery.*,
    sequence.Description,
    song.Artist, song.Title,
    dancer.Name
    FROM mastery
    INNER JOIN sequence ON sequence.SequenceID=mastery.SequenceID
    INNER JOIN song ON sequence.SongID=song.SongID
    INNER JOIN dancer ON dancer.DancerID=mastery.DancerID
    """)
    data = res.fetchall()
    con.close()

    answer = []
    for entry in data:
        answer.append(
            {
                "ID": entry[0],
                "SequenceID": entry[1],
                "DancerID": entry[2],
                "Description": entry[3],
                "Artist": entry[4],
                "Title": entry[5],
                "Name": entry[6],
            }
        )
    return answer

@app.get("/sequence")
def read_root():
    con = sqlite3.connect("songlist.db")
    cur = con.cursor()
    res = cur.execute("""
    SELECT 
    sequence.*,
    song.Artist, song.Title
    FROM sequence
    INNER JOIN song ON sequence.SongID=song.SongID
    """)
    data = res.fetchall()
    con.close()

    answer = []
    for entry in data:
        answer.append(
            {
                "ID": entry[0],
                "SongID": entry[1],
                "Start": entry[2],
                "End": entry[3],
                "Description": entry[4],
                "Comment": entry[5],
                "LastTimeInRDG": entry[6],
                "Artist": entry[7],
                "Title": entry[8],
            }
        )
    return answer

@app.get("/song")
def read_root():
    con = sqlite3.connect("songlist.db")
    cur = con.cursor()
    res = cur.execute("""
    SELECT 
    *
    FROM song
    """)
    data = res.fetchall()
    con.close()

    answer = []
    for entry in data:
        answer.append(
            {
                "ID": entry[0],
                "URL": entry[1],
                "Artist": entry[2],
                "Title": entry[3],
                "Comment": entry[4],
            }
        )
    return answer

@app.get("/dancer")
def read_root():
    con = sqlite3.connect("songlist.db")
    cur = con.cursor()
    res = cur.execute("""
    SELECT 
    *
    FROM dancer
    """)
    data = res.fetchall()
    con.close()

    answer = []
    for entry in data:
        answer.append(
            {
                "ID": entry[0],
                "Name": entry[1],
            }
        )
    return answer

@app.get("/all")
def read_root():
    con = sqlite3.connect("songlist.db")
    cur = con.cursor()
    res = cur.execute("""
    SELECT 
    song.URL, song.Title, song.Artist, 
    sequence.Description, sequence.Start, sequence.End, sequence.LastTimeInRDG, sequence.comment, 
    dancer.Name
    FROM sequence
    INNER JOIN song ON sequence.SongID=song.SongID
    INNER JOIN mastery ON sequence.SequenceID=mastery.SequenceID
    INNER JOIN dancer ON mastery.DancerID=dancer.DancerID
    """)
    data = res.fetchall()
    con.close()

    answer = []
    for entry in data:
        answer.append(
            {
                "URL": entry[0],
                "Title": entry[1],
                "Artist": entry[2],
                "Description": entry[3],
                "Start": entry[4],
                "End": entry[5],
                "LastTimeInRDG": entry[6],
                "Comment": entry[7],
                "Dancer": entry[8],
            }
        )
    return answer
